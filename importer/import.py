"""
xlsx -> PostgreSQL importer.
Reads the exported Google Sheet (cached values) and generates import_data.sql,
which is then applied with psql. Run:
    python import.py
    psql -U postgres -d bookkeeping -f import_data.sql
"""
import openpyxl, datetime, os, re

SRC = r"C:\Users\PC\Downloads\chuakatherinemayroluna@gmail.com.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "import_data.sql")

wb = openpyxl.load_workbook(SRC, data_only=True)

def quote(s):
    return "'" + s.replace("'", "''") + "'"

def esc(v):
    if v is None or v == "":
        return "NULL"
    if isinstance(v, (int, float)):
        return repr(round(float(v), 4))
    if isinstance(v, datetime.datetime):
        return f"'{v.date().isoformat()}'"
    if isinstance(v, datetime.date):
        return f"'{v.isoformat()}'"
    s = str(v).strip()
    if not s or s in ("-", "#N/A", "#REF!", "#VALUE!", "COMPUTED_VALUE"):
        return "NULL"
    return "'" + s.replace("'", "''") + "'"

def num(v):
    """Numeric or 0."""
    if isinstance(v, (int, float)):
        return repr(round(float(v), 4))
    if isinstance(v, str):
        s = re.sub(r"[^0-9.\-]", "", v)
        try:
            return repr(float(s))
        except ValueError:
            return "0"
    return "0"

def is_date(v):
    return isinstance(v, (datetime.date, datetime.datetime))

def sale_no(v):
    """SALES # as clean text: 4.0 -> '4'."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

sql = ["BEGIN;", "SET client_min_messages TO WARNING;"]

# ---------- settings (Setup tab) ----------
setup = wb["Setup"]
currency = setup["C9"].value or "PHP"
sql.append(f"INSERT INTO settings (key,value) VALUES ('currency_symbol',{esc(currency)}) ON CONFLICT (key) DO NOTHING;")

# ---------- vendors ----------
ws = wb["Vendors"]
n_vend = 0
for r in range(7, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    sql.append(
        "INSERT INTO vendors (name,contact_name,phone,email,address,country,notes) VALUES ("
        + ",".join(esc(ws.cell(r, c).value) for c in (2, 3, 4, 5, 6, 7, 8))
        + ") ON CONFLICT (name) DO NOTHING;")
    n_vend += 1

# ---------- accounts ----------
ws = wb["Accounts"]
n_acct = 0
for r in range(15, 55):
    name = ws.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    beg = ws.cell(r, 6).value  # F = BEGINNING BALANCE
    sql.append(
        f"INSERT INTO accounts (name,beginning_balance) VALUES ({esc(name)},{num(beg)}) "
        "ON CONFLICT (name) DO NOTHING;")
    n_acct += 1

# ---------- items (Inventory tab, rows 9+) ----------
ws = wb["Inventory"]
n_items = 0
for r in range(9, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    typ = str(ws.cell(r, 5).value or "Product").strip()
    if typ not in ("Product", "Material"):
        typ = "Product"
    vend = ws.cell(r, 14).value  # N = preferred vendor (by name)
    sql.append(
        "INSERT INTO items (name,sku,category,type,initial_stock,minimum_stock,"
        "sales_price,cost,units_in_purchase,promotion,notes,preferred_vendor_id) VALUES ("
        f"{esc(name)},{esc(ws.cell(r,3).value)},{esc(ws.cell(r,4).value)},'{typ}',"
        f"{num(ws.cell(r,6).value)},{num(ws.cell(r,8).value)},"
        f"{num(ws.cell(r,10).value)},{num(ws.cell(r,11).value)},"
        f"{num(ws.cell(r,15).value)},{esc(ws.cell(r,16).value)},{esc(ws.cell(r,17).value)},"
        f"(SELECT id FROM vendors WHERE name={esc(vend)})"
        ") ON CONFLICT (name) DO NOTHING;")
    n_items += 1

# ---------- sales reps (distinct SALES REP values in Sales Database col P) ----------
ws = wb["Sales Database"]
reps = set()
for r in range(6, ws.max_row + 1):
    rep = ws.cell(r, 16).value
    if rep and str(rep).strip() and str(rep).strip() != "-":
        reps.add(str(rep).strip())
for rep in sorted(reps):
    sql.append(f"INSERT INTO sales_reps (name) VALUES ({esc(rep)}) ON CONFLICT (name) DO NOTHING;")

# ---------- sales + line items ----------
# Multi-row invoices: a row with SALES # starts an invoice; following rows with
# empty A but an ITEM continue the same invoice's line items.
n_sales = 0
n_lines = 0
cur = None
def flush(cur):
    global n_sales, n_lines
    if not cur:
        return
    s = cur
    sql.append(
        "INSERT INTO sales (sales_no,date,customer,store_farm,term,contact_no,payment_mode,"
        "account_id,sales_rep_id,subtotal,tax_pct,tax_amount,discount_pct,discount,total,amount_paid,status) VALUES ("
        f"{esc(s['no'])},{esc(s['date'])},{esc(s['customer'])},{esc(s['store'])},{esc(s['term'])},"
        f"{esc(s['contact'])},{esc(s['mode'])},"
        f"(SELECT id FROM accounts WHERE name={esc(s['mode'])}),"
        f"(SELECT id FROM sales_reps WHERE name={esc(s['rep'])}),"
        f"{num(s['subtotal'])},{num(s['taxpct'])},"
        f"{num(s['subtotal'])}*{num(s['taxpct'])}/100,"
        f"{num(s['discpct'])},{num(s['disc'])},{num(s['total'])},"
        # Cash sales treated as fully paid; credit terms start unpaid
        f"CASE WHEN {esc(s['term'])} IS NULL OR {esc(s['term'])} ILIKE '%cash%' THEN {num(s['total'])} ELSE 0 END,"
        "'Completed') ON CONFLICT (sales_no) DO NOTHING;")
    n_sales += 1
    for (item, qty, price) in s["lines"]:
        sql.append(
            "INSERT INTO sale_items (sale_id,item_id,qty,unit_price,total_price) "
            f"SELECT s.id, i.id, {num(qty)}, "
            f"CASE WHEN {num(qty)}=0 THEN 0 ELSE {num(price)}/{num(qty)} END, {num(price)} "
            f"FROM sales s, items i WHERE s.sales_no={esc(s['no'])} AND i.name={esc(item)};")
        n_lines += 1

for r in range(6, ws.max_row + 1):
    a = ws.cell(r, 1).value
    item = ws.cell(r, 8).value
    if a not in (None, "") and is_date(ws.cell(r, 2).value) and (cur is None or sale_no(a) != cur["no"]):
        flush(cur)
        cur = dict(no=sale_no(a), date=ws.cell(r, 2).value, customer=ws.cell(r, 3).value,
                   store=ws.cell(r, 4).value, term=ws.cell(r, 5).value,
                   contact=ws.cell(r, 6).value, mode=ws.cell(r, 7).value,
                   subtotal=ws.cell(r, 11).value, taxpct=ws.cell(r, 12).value,
                   discpct=ws.cell(r, 13).value, disc=ws.cell(r, 14).value,
                   total=ws.cell(r, 15).value, rep=ws.cell(r, 16).value, lines=[])
        if item and str(item).strip():
            cur["lines"].append((item, ws.cell(r, 9).value, ws.cell(r, 10).value))
    elif cur and item and str(item).strip():
        cur["lines"].append((item, ws.cell(r, 9).value, ws.cell(r, 10).value))
flush(cur)

# ---------- purchases ("Sold Items" tab, rows 7+) ----------
ws = wb["Sold Items"]
n_po = 0
for r in range(7, ws.max_row + 1):
    if not is_date(ws.cell(r, 2).value):
        continue
    item = ws.cell(r, 5).value
    if not item or not str(item).strip():
        continue
    sql.append(
        "INSERT INTO purchases (order_date,received_date,ref_id,item_id,purchase_qty,"
        "received_qty,unit_cost,account_id,status,vendor_id,notes) VALUES ("
        f"{esc(ws.cell(r,2).value)},{esc(ws.cell(r,3).value) if is_date(ws.cell(r,3).value) else 'NULL'},"
        f"{esc(ws.cell(r,4).value)},"
        f"(SELECT id FROM items WHERE name={esc(item)}),"
        f"{num(ws.cell(r,7).value)},{num(ws.cell(r,8).value)},{num(ws.cell(r,10).value)},"
        f"(SELECT id FROM accounts WHERE name={esc(ws.cell(r,12).value)}),"
        f"{esc(ws.cell(r,13).value) if ws.cell(r,13).value else quote('Ordered')},"
        f"(SELECT id FROM vendors WHERE name={esc(ws.cell(r,14).value)}),"
        f"{esc(ws.cell(r,16).value)});")
    n_po += 1

# ---------- expenses (rows 7+) ----------
ws = wb["Expenses"]
n_exp = 0
for r in range(7, ws.max_row + 1):
    if not is_date(ws.cell(r, 2).value):
        continue
    sql.append(
        "INSERT INTO expenses (date,ref_id,category,amount,tax,shipping,fees,account_id,description,remarks) VALUES ("
        f"{esc(ws.cell(r,2).value)},{esc(ws.cell(r,3).value)},{esc(ws.cell(r,4).value) or quote('Uncategorized')},"
        f"{num(ws.cell(r,5).value)},{num(ws.cell(r,6).value)},{num(ws.cell(r,7).value)},{num(ws.cell(r,8).value)},"
        f"(SELECT id FROM accounts WHERE name={esc(ws.cell(r,10).value)}),"
        f"{esc(ws.cell(r,11).value)},{esc(ws.cell(r,12).value)});")
    n_exp += 1

# ---------- balance entries / manual inventory ----------
for sheet, table, cols in (
    ("Balance", "balance_entries", None),
    ("Manual Inventory", "manual_inventory", None),
):
    ws = wb[sheet]
    for r in range(7, ws.max_row + 1):
        if not is_date(ws.cell(r, 2).value):
            continue
        if table == "balance_entries":
            sql.append(
                "INSERT INTO balance_entries (date,ref_id,account_id,amount,description,remarks) VALUES ("
                f"{esc(ws.cell(r,2).value)},{esc(ws.cell(r,3).value)},"
                f"(SELECT id FROM accounts WHERE name={esc(ws.cell(r,4).value)}),"
                f"{num(ws.cell(r,5).value)},{esc(ws.cell(r,6).value)},{esc(ws.cell(r,7).value)});")
        else:
            sql.append(
                "INSERT INTO manual_inventory (date,batch_no,item_id,qty,notes) VALUES ("
                f"{esc(ws.cell(r,2).value)},{esc(ws.cell(r,3).value)},"
                f"(SELECT id FROM items WHERE name={esc(ws.cell(r,4).value)}),"
                f"{num(ws.cell(r,6).value)},{esc(ws.cell(r,7).value)});")

sql.append("COMMIT;")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))

print(f"vendors={n_vend} accounts={n_acct} items={n_items} reps={len(reps)} "
      f"sales={n_sales} lines={n_lines} purchases={n_po} expenses={n_exp}")
print("Wrote", OUT)
