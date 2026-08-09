"""
MASTERLIST import — "Elishen URC Sales until Jul 20.xlsx".
- Groups continuation rows into invoices (date/customer/discount/total on first row).
- EXCLUDES Jerome Racaza and Fegen Gines (payment still being tracked, per owner).
- PAID -> payments ledger row (account unknown -> NULL); UNPAID/blank -> receivable.
- Inserts stock-neutral offset batches: these sales predate the 7/29 stock take,
  which already reflects them — without offsets stock would double-deduct.
Generates masterlist.sql; prints unmatched product names for review.
"""
import openpyxl, re, os, datetime
from collections import defaultdict

SRC = r"C:\Users\PC\Downloads\Elishen URC Sales until Jul 20.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "masterlist.sql")
EXCLUDE = re.compile(r"jerome\s*racaza|fegen\s*gines", re.I)

ws = openpyxl.load_workbook(SRC, data_only=True)["MASTERLIST"]

def norm(s): return re.sub(r"\s+", " ", str(s)).strip().lower()
def esc(s): return str(s).replace("'", "''")

def qty_of(v):
    """'20 (+1)'->21, '72 (60+12)'->72, '12 (10+2)'->12, '6 packs'->6, '1 box'->1"""
    s = norm(v)
    m = re.match(r"^(\d+(?:\.\d+)?)", s)
    lead = float(m.group(1)) if m else 1.0
    plus = re.search(r"\(\s*\+\s*(\d+)\s*\)", s)          # "(+1)" = extra free on top
    if plus: return lead + float(plus.group(1))
    return lead

def num(v):
    try: return round(float(v), 2)
    except (TypeError, ValueError): return None

# ---- product-name mapping (masterlist shorthand -> DB item name) ----
def map_item(raw):
    s = norm(raw)
    def has(*ws_): return all(w in s for w in ws_)
    # Supremo Infinity
    if s.startswith(("si ", "rm", "fortifier", "conditioning", "super", "power", "si")) or "si " in s or s.startswith("rm ("):
        if has("super conditioner"): return "Supremo Infinity Super Conditioner (25KG)"
        if has("power"): return "Supremo Infinity Power Concentrate (25KG)"
        if "fortifier" in s or has("32"): return "Supremo Infinity 32 Pellets - 32% CP (25x1kg)"
        if "23" in s or "conditioning" in s: return "Supremo Infinity 23 Conditioning (25x1kg)"
        if "chick booster" in s: return "Supremo Infinity 1 Booster (25x1kg)"
        if "rm" in s:
            if "(50" in s: return "Supremo Infinity Ready Mix - Grains + Pellets (RED) (50kg)"
            if "(25x1" in s: return "Supremo Infinity Ready Mix red (25x1kg)"
            if "(25kg" in s or "(25 kg" in s: return "Supremo Infinity Ready Mix - Grains + Pellets (RED) (25x1kg)"
            return "Supremo Infinity Ready Mix red (25x1kg)"   # 'SI RM' pack
        if "2.1" in s:
            return "Supremo Infinity 2.1 - Developer - 3 Grains (50KG)" if "(50" in s \
              else "Supremo Infinity 2.1 Developer + (25x1kg)"
        if re.search(r"\b1\b", s):
            return "Supremo Infinity 1 - Chick Booster Crumble (50KG)" if "(50" in s \
              else "Supremo Infinity 1 Booster (25x1kg)"
        if re.search(r"\b2\b", s):
            return "Supremo Infinity 2 - Chick Grower Crumble (50KG)" if "(50" in s \
              else "Supremo Infinity 2 Grower (25x1kg)"
        if re.search(r"\b3\b", s):
            return "Supremo Infinity 3 - Maintenance Pellets - 15% CP (50KG)"
        if re.search(r"\b4\b", s):
            return "Supremo Infinity 4 - Breeder Pellets (50KG)" if "(50" in s \
              else "Supremo Infinity 4 Breeder (25x1kg)"
    # UNO+ / Supreme / Stargain
    if "supreme lactating" in s: return "UNO+ Supreme Lactating (50KG)"
    if s.startswith("uno"):
        if "pre" in s: return "UNO+ Pre Starter (25KG)"
        if "boost" in s: return "UNO+ Booster (25x1kg)"
        if "starter" in s: return "UNO+ Starter (50KG)"
        if "grower" in s: return "UNO+ Grower (50KG)"
        if "finisher" in s: return "UNO+ Finisher (50KG)"
        if "breeder" in s: return "UNO+ Breeder (50KG)"
        if "lactating" in s: return "UNO+ Lactating (50KG)"
    if s.startswith("sg"):
        if "starter" in s: return "Stargain Starter (50KG)"
        if "grower" in s: return "Stargain Grower (50KG)"
        if "breeder" in s: return "Stargain Breeder (50KG)"
        if "finisher" in s: return "Stargain Finisher (50KG)"
        if "lactating" in s: return "Stargain Lactating (50KG)"
    # Topbreed
    if s.startswith("tb") or "topbreed" in s:
        five = "(5kg" in s or "(5 kg" in s
        if "cat" in s: return "Topbreed Cat Adult (5KG)" if five else "Topbreed Cat Adult (20KG)"
        if "puppy" in s: return "Topbreed Dog Puppy (20KG)"
        if "mini" in s: return "Topbreed Dog Adult Mini (5KG)" if five else "Topbreed Dog Adult Mini (20KG)"
        if "adult" in s: return "Topbreed Dog Adult (5KG)" if five else "Topbreed Dog Adult (20KG)"
    # Cat litter
    if s.startswith("cl") or "lavender" in s or "coffee" in s:
        if "lavender" in s: return "Top Care CAT LITTER Lavander (10L)"
        if "coffee" in s: return "Top Care CAT LITTER Coffee (10L)"
    # Supplies / medicines
    if "top b" in s or "tb+ vitamins" in s: return "Top B+ Vitamins 60ml"
    if "shampooch" in s:
        return "Shampooch 300ml bt 12/box" if "300" in s else "Shampooch Sachet"
    if "iron d" in s or "iron-d" in s: return "Iron - D inj"
    if "robicomject" in s: return "Robicomject inj 100ml"
    if "tripulac" in s: return "Tripulac Pig Doser 2x1 set"
    if "wheat germ" in s or "wheatgerm" in s: return "Wheatgerm 300Gm x 12/box"
    if "robipenstr" in s: return "Robipenstrep P 10dose/Diluent"
    return None

# ---- parse invoices ----
invoices, cur, unmatched = [], None, set()
for r in range(2, ws.max_row + 1):
    date, cust = ws.cell(r, 1).value, ws.cell(r, 2).value
    prod, qty, amt = ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value
    if isinstance(date, datetime.datetime) and cust:
        if cur: invoices.append(cur)
        disc_raw = ws.cell(r, 7).value
        disc = num(disc_raw) if not isinstance(disc_raw, str) or not disc_raw.strip() or disc_raw.strip()[0].isdigit() else 0
        cur = {"date": date.date().isoformat(), "customer": re.sub(r"\s+", " ", str(cust)).strip(),
               "address": re.sub(r"\s+", " ", str(ws.cell(r, 3).value)).strip() if ws.cell(r, 3).value else None,
               "discount": disc or 0, "total": num(ws.cell(r, 8).value),
               "status": norm(ws.cell(r, 9).value or ""), "lines": []}
    if cur is not None and prod and str(prod).strip():
        name = map_item(prod)
        if name is None: unmatched.add(str(prod).strip())
        cur["lines"].append({"raw": str(prod).strip(), "item": name,
                             "qty": qty_of(qty) if qty is not None else 1.0, "amount": num(amt)})
if cur: invoices.append(cur)

kept = [i for i in invoices if not EXCLUDE.search(i["customer"])]
skipped = [i for i in invoices if EXCLUDE.search(i["customer"])]
paid = [i for i in kept if i["status"] == "paid"]

if unmatched:
    print("UNMATCHED PRODUCTS:")
    for u in sorted(unmatched): print("  -", u)

# ---- SQL ----
sql = ["BEGIN;"]
offset_qty = defaultdict(float)
for n, inv in enumerate(kept, 1):
    sn = f"ML-{n:03d}"
    line_sum = sum(l["amount"] or 0 for l in inv["lines"])
    total = inv["total"] if inv["total"] is not None else max(line_sum - inv["discount"], 0)
    subtotal = line_sum if line_sum > 0 else total + inv["discount"]
    paid_flag = inv["status"] == "paid"
    sql.append(
        "INSERT INTO sales (sales_no, date, customer, store_farm, subtotal, tax_pct, tax_amount, "
        "discount_pct, discount, total, amount_paid, status, term) VALUES ("
        f"'{sn}', '{inv['date']}', '{esc(inv['customer'])}', "
        f"{('null' if not inv['address'] else chr(39)+esc(inv['address'])+chr(39))}, "
        f"{round(subtotal,2)}, 0, 0, 0, {round(inv['discount'],2)}, {round(total,2)}, "
        f"{round(total,2) if paid_flag else 0}, 'Completed', "
        f"{'null' if paid_flag else chr(39)+'Credit'+chr(39)});")
    for l in inv["lines"]:
        if l["item"] is None: continue
        qty = l["qty"] or 1
        amount = l["amount"] if l["amount"] is not None else 0
        up = round(amount / qty, 4) if qty else 0
        sql.append(
            "INSERT INTO sale_items (sale_id, item_id, qty, unit_price, total_price) "
            f"SELECT s.id, i.id, {qty}, {up}, {amount} FROM sales s, items i "
            f"WHERE s.sales_no = '{sn}' AND i.name = '{esc(l['item'])}';")
        offset_qty[l["item"]] += qty
    if paid_flag:
        sql.append(
            "INSERT INTO payments (sale_id, date, amount, account_id, notes) "
            f"SELECT s.id, '{inv['date']}', {round(total,2)}, NULL, "
            f"'Masterlist import — marked PAID (account not recorded)' "
            f"FROM sales s WHERE s.sales_no = '{sn}';")

# stock-neutral offsets: 7/29 stock take already includes these historical sales
for item, q in sorted(offset_qty.items()):
    sql.append(
        "INSERT INTO manual_inventory (date, batch_no, item_id, qty, notes) "
        f"SELECT '2026-07-29', 'PRE-STOCKTAKE-OFFSET', i.id, {round(q,3)}, "
        f"'Offset: masterlist sales predate the 7/29 physical count' "
        f"FROM items i WHERE i.name = '{esc(item)}';")

sql.append("COMMIT;")
sql.append("SELECT 'invoices: '||count(*) FROM sales WHERE sales_no LIKE 'ML-%';")
sql.append("SELECT 'income: '||SUM(total)||' | paid: '||SUM(amount_paid) FROM sales WHERE sales_no LIKE 'ML-%';")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))
print(f"invoices kept: {len(kept)} (paid {len(paid)}, unpaid/blank {len(kept)-len(paid)})")
print(f"EXCLUDED (Racaza/Gines, payment tracking): {len(skipped)} invoices, "
      f"total {sum(i['total'] or 0 for i in skipped):,.2f}")
print("wrote", OUT)
