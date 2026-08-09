"""
FULL-WIDTH extraction of all four sheets of "URC net Capital .xlsx".
Adds the dealer-side economics the earlier passes missed:
  - dealer build-up (Add block): income, FTH, dist-to-dealer, funds, dealer/cash
    discounts, Ktech incentives -> NET dealer price
  - published SRP per BAG and per KILO (feeds/pets)
  - RobiChem: DEALERS ACQUISITION, rounded dealer SRP, distributor profit, margin
  - Pets DEALS column ("20 + 1 bag")
Generates full_matrix.sql. Values are taken verbatim from sheet cells.
"""
import openpyxl, json, os, re

SRC = r"C:\Users\PC\Downloads\URC net Capital .xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "full_matrix.sql")
wb = openpyxl.load_workbook(SRC, data_only=True)

def num(v, nd=4):
    try: return round(float(v), nd)
    except (TypeError, ValueError): return None

def clean(s): return re.sub(r"\s+", " ", str(s)).strip()
def esc(s): return str(s).replace("'", "''")

sql = ["BEGIN;",
"""ALTER TABLE items
  ADD COLUMN IF NOT EXISTS srp_kilo numeric(14,4),
  ADD COLUMN IF NOT EXISTS dealer_srp numeric(14,4),
  ADD COLUMN IF NOT EXISTS dealers_acquisition numeric(14,4);"""]

def upd(name, cols, bdpatch):
    sets = [f"{k} = {v if v is not None else 'NULL'}" for k, v in cols.items()]
    if bdpatch:
        bd = json.dumps(bdpatch).replace("'", "''")
        sets.append(f"price_breakdown = COALESCE(price_breakdown, '{{}}'::jsonb) || '{bd}'::jsonb")
    sql.append(f"UPDATE items SET {', '.join(sets)} WHERE name = '{esc(name)}';")

# ---------- GameFowl: J cap K income L fth M dist N sales O mp P bd Q dealer R cash S ktech T NET V bag W kilo
GF = ["Supremo Infinity 1 - Chick Booster Crumble (50KG)",
      "Supremo Infinity 2 - Chick Grower Crumble (50KG)",
      "Supremo Infinity 2.1 - Developer - 3 Grains (50KG)",
      "Supremo Infinity 3 - Maintenance Pellets - 15% CP (50KG)",
      "Supremo Infinity 3 - Maintenance Pellets - 15% CP w/ Grain (50KG)",
      "Supremo Infinity 4 - Breeder Pellets (50KG)",
      "Supremo Infinity Ready Mix - Grains + Pellets (RED) (50kg)",
      "Supremo Infinity Ready Mix - Grains + Pellets (RED) (25x1kg)",
      "Supremo Infinity 1 Booster (25x1kg)",
      "Supremo Infinity 2 Grower (25x1kg)",
      "Supremo Infinity 2.1 Developer + (25x1kg)",
      "Supremo Infinity 4 Breeder (25x1kg)",
      "Supremo Infinity Ready Mix red (25x1kg)",
      "Supremo Infinity 23 Conditioning (25x1kg)",
      "Supremo Infinity 32 Pellets - 32% CP (25x1kg)",
      "Supremo Infinity Power Concentrate (25KG)",
      "Supremo Infinity Super Conditioner (25KG)"]
ws = wb["GameFowl"]
for i, name in enumerate(GF):
    r = 7 + i
    build = {"distributor_income": num(ws.cell(r, 11).value), "fth": num(ws.cell(r, 12).value),
             "dist_to_dealer": num(ws.cell(r, 13).value), "sales_fund": num(ws.cell(r, 14).value),
             "manpower_fund": num(ws.cell(r, 15).value), "bus_devt": num(ws.cell(r, 16).value),
             "dealer_discount": num(ws.cell(r, 17).value), "cash_discount": num(ws.cell(r, 18).value),
             "ktech": num(ws.cell(r, 19).value), "net_dealer_price": num(ws.cell(r, 20).value)}
    upd(name, {"dealers_acquisition": num(ws.cell(r, 20).value),
               "sales_price": num(ws.cell(r, 22).value),
               "srp_kilo": num(ws.cell(r, 23).value)}, {"dealer_build": build})

# ---------- Hogs: J cap L income M plant-whse N whse-dealer O sales P mp Q bd R outright S cod T ktech U bag W pub-bag X pub-kilo
HG = [(7, "UNO+ Booster (25x1kg)"), (8, "UNO+ Supreme Pre Starter Crumble (25KG)"),
      (9, "UNO+ Supreme Starter Pellet (50KG)"), (10, "UNO+ Supreme Grower (50KG)"),
      (11, "UNO+ Supreme Finisher (50KG)"), (12, "UNO+ Supreme Breeder (50KG)"),
      (13, "UNO+ Supreme Lactating (50KG)"), (16, "UNO+ Pre Starter (25KG)"),
      (17, "UNO+ Starter (50KG)"), (18, "UNO+ Grower (50KG)"), (19, "UNO+ Finisher (50KG)"),
      (20, "UNO+ Breeder (50KG)"), (21, "UNO+ Lactating (50KG)"),
      (27, "Stargain Starter (50KG)"), (28, "Stargain Grower (50KG)"),
      (29, "Stargain Finisher (50KG)"), (30, "Stargain Breeder (50KG)"),
      (31, "Stargain Lactating (50KG)")]
ws = wb["Hogs"]
for r, name in HG:
    build = {"distributor_income": num(ws.cell(r, 12).value), "fth": num(ws.cell(r, 13).value),
             "dist_to_dealer": num(ws.cell(r, 14).value), "sales_fund": num(ws.cell(r, 15).value),
             "manpower_fund": num(ws.cell(r, 16).value), "bus_devt": num(ws.cell(r, 17).value),
             "dealer_discount": num(ws.cell(r, 18).value), "cash_discount": num(ws.cell(r, 19).value),
             "ktech": num(ws.cell(r, 20).value), "net_dealer_price": num(ws.cell(r, 21).value)}
    upd(name, {"dealers_acquisition": num(ws.cell(r, 21).value),
               "sales_price": num(ws.cell(r, 23).value),
               "srp_kilo": num(ws.cell(r, 24).value)}, {"dealer_build": build})

# ---------- Pets: L cap M income N fth O dist P sales Q mp R bd S dealer T cash U ktech V NET W tactical Y bag Z kilo AD deals
PT = [(6, "Topbreed Dog Puppy (20KG)"), (7, "Topbreed Dog Adult (20KG)"),
      (8, "Topbreed Dog Adult Mini (20KG)"), (9, "Topbreed Cat Adult (20KG)"),
      (10, "Topbreed Dog Adult (5KG)"), (11, "Topbreed Dog Adult Mini (5KG)"),
      (12, "Topbreed Cat Adult (5KG)"), (13, "Topbreed Dog Puppy 2kgx10"),
      (14, "Gravy Chunks (Roasted Chicken&Liver/Chicken&Liver Steak) 130gx12x4"),
      (15, "Top Care CAT LITTER (10Lx3)")]
ws = wb["Pets"]
for r, name in PT:
    deal_raw = ws.cell(r, 30).value
    deal = clean(deal_raw) if isinstance(deal_raw, str) and "+" in str(deal_raw) else None
    build = {"distributor_income": num(ws.cell(r, 13).value), "fth": num(ws.cell(r, 14).value),
             "dist_to_dealer": num(ws.cell(r, 15).value), "sales_fund": num(ws.cell(r, 16).value),
             "manpower_fund": num(ws.cell(r, 17).value), "bus_devt": num(ws.cell(r, 18).value),
             "dealer_discount": num(ws.cell(r, 19).value), "cash_discount": num(ws.cell(r, 20).value),
             "ktech": num(ws.cell(r, 21).value), "net_dealer_price": num(ws.cell(r, 22).value),
             "tactical_fund": num(ws.cell(r, 23).value)}
    cols = {"dealers_acquisition": num(ws.cell(r, 22).value),
            "sales_price": num(ws.cell(r, 25).value),
            "srp_kilo": num(ws.cell(r, 26).value)}
    if deal:
        sql.append(f"UPDATE items SET deal = '{esc(deal)}' WHERE name = '{esc(name)}';")
    upd(name, cols, {"dealer_build": build})

# ---------- RobiChem: N acquisition O dealer SRP P profit Q margin (sections share columns)
EXISTING = {
    ("COCCIBUSTER", "5g."): "Coccibuster", ("LEVOMAX", "5g."): "Levomax",
    ("SPECTRUM", "5g."): "Spectrum (96/box)", ("SPECTRUM PLUS", "5g."): "Spectrum Plus (96/box)",
    ("ROBISTREP VK Powder", "5g."): "Robistrep Vk 5Gm",
    ("WORM BUSTER", "5g."): "Wormbuster Single Dose 5Gm",
    ("ROBI LA Inj.", "100 ml"): "Robi L.A inj 100ml",
    ("ROBIPENSTREP P", "10 ds w/ diluent"): "Robipenstrep P 10dose/Diluent",
    ("ROBIPENSTREP P", "Single dose"): "Robipenstrep P Single dose bt",
    ("IRON - D Inj.", "100 ml"): "Iron - D inj",
    ("ROBICOMJECT Injectible", "100 ml"): "Robicomject inj 100ml",
    ("TRIPULAC Pig Doser", "2 x 100ml (disp.)"): "Tripulac Pig Doser 2x1 set",
    ("WHEAT GERM /BOX", "300g"): "Wheatgerm 300Gm x 12/box",
}
MERGED = {  # dealer-section rows live on the stocked Pet Supplies items now
    ("TopB + Multivitamins", "120ml x 18bottles"): "Top B+ Vitamins 120ml",
    ("TopB + Multivitamins", "60ml x 36bottles"): "Top B+ Vitamins 60ml",
    ("Shampooch", "300ml x 12bottles"): "Shampooch 300ml bt 12/box",
    ("Shampooch", "15ml x 25sachets x 16box"): "Shampooch Sachet",
}
def title_name(product, pack):
    p = re.sub(r"\s+", " ", product).strip().title()
    p = p.replace("Scm", "SCM").replace("Vk", "VK").replace("La", "LA").replace("Ii", "II")
    return f"{p} {pack}"
def db_name(product, pack):
    for m in (EXISTING, MERGED):
        for (p, k), n in m.items():
            if clean(p).lower() == product.lower() and clean(k).lower() == pack.lower():
                return n
    return title_name(product, pack)

ws = wb["RobiChem"]
current = None
for r in range(4, ws.max_row + 1):
    p, pack = ws.cell(r, 1).value, ws.cell(r, 2).value
    inv = num(ws.cell(r, 3).value)
    if p and clean(p): current = clean(p)
    if inv is None or pack is None or not current: continue
    acq, dsrp = num(ws.cell(r, 14).value), num(ws.cell(r, 15).value)
    profit, margin = num(ws.cell(r, 16).value), num(ws.cell(r, 17).value)
    if acq is None and dsrp is None: continue
    name = db_name(current, clean(pack))
    upd(name, {"dealers_acquisition": acq, "dealer_srp": dsrp},
        {"distributor_profit": profit, "margin_rate": margin})

sql.append("COMMIT;")
sql.append("""SELECT name, dealers_acquisition, dealer_srp, sales_price, srp_kilo
FROM items WHERE name IN ('Supremo Infinity 1 - Chick Booster Crumble (50KG)',
 'UNO+ Booster (25x1kg)','Coccibuster','Topbreed Dog Adult (5KG)','Top Care CAT LITTER (10Lx3)',
 'Shampooch Sachet') ORDER BY name;""")
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))
print("wrote", OUT)
