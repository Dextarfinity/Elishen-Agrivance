"""
Faithful re-extraction of the RobiChem sheet — ALL FOUR SECTIONS:
  R4-34   standard: LESS 20% / 10% / 20% + 12% VAT, SRP x1.4, deals + dealer deals (col M)
  R37-40  Wheat Germ: 10% / 5% / 20% + pick-up pesos, NO VAT
  R45-58  SCM: 10% volume / 20% pick-up / 5% PBD + 12% VAT (no SRP in file, except Robophos)
  R63-67  Dealers Discount: peso Volume/PD/PBD + 12% VAT, SRP, outright/COD
Every capital/SRP is taken from the sheet's own computed cells — nothing re-derived.
Also merges duplicate TopB+/Shampooch rows into the stocked Pet Supplies items.
"""
import openpyxl, json, os, re

SRC = r"C:\Users\PC\Downloads\URC net Capital .xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fix_robichem.sql")
ws = openpyxl.load_workbook(SRC, data_only=True)["RobiChem"]

def num(v, nd=4):
    try: return round(float(v), nd)
    except (TypeError, ValueError): return None

def clean(s): return re.sub(r"\s+", " ", str(s)).strip()

def title_name(product, pack):
    p = re.sub(r"\s+", " ", product).strip().title()
    p = p.replace("Scm", "SCM").replace("Vk", "VK").replace("La", "LA").replace("Ii", "II")
    return f"{p} {pack}"

# names created by the earlier import for each (product, packaging) — for standard rows
EXISTING = {
    ("COCCIBUSTER", "5g."): "Coccibuster",
    ("LEVOMAX", "5g."): "Levomax",
    ("SPECTRUM", "5g."): "Spectrum (96/box)",
    ("SPECTRUM PLUS", "5g."): "Spectrum Plus (96/box)",
    ("ROBISTREP VK Powder", "5g."): "Robistrep Vk 5Gm",
    ("WORM BUSTER", "5g."): "Wormbuster Single Dose 5Gm",
    ("ROBI LA Inj.", "100 ml"): "Robi L.A inj 100ml",
    ("ROBIPENSTREP P", "10 ds w/ diluent"): "Robipenstrep P 10dose/Diluent",
    ("ROBIPENSTREP P", "Single dose"): "Robipenstrep P Single dose bt",
    ("IRON - D Inj.", "100 ml"): "Iron - D inj",
    ("ROBICOMJECT Injectible", "100 ml"): "Robicomject inj 100ml",
    ("TRIPULAC Pig Doser", "2 x 100ml (disp.)"): "Tripulac Pig Doser 2x1 set",
    ("WHEAT GERM /BOX", "300g"): "Wheatgerm 300Gm x 12/box",
}
def db_name(product, pack):
    for (p, k), n in EXISTING.items():
        if clean(p).lower() == product.lower() and clean(k).lower() == pack.lower():
            return n
    return title_name(product, pack)

def esc(s): return str(s).replace("'", "''")

sql = ["BEGIN;"]
updates = []

def upsert(name, capital, srp, deal, outright, cod, breakdown, packaging):
    bd = json.dumps(breakdown).replace("'", "''")
    srp_sql = srp if srp is not None else "NULL"
    deal_sql = f"'{esc(deal)}'" if deal else "NULL"
    sql.append(
        f"UPDATE items SET cost = {capital}, sales_price = {srp_sql}, deal = {deal_sql}, "
        f"outright_rate = {outright}, cod_rate = {cod}, packaging = '{esc(packaging)}', "
        f"price_breakdown = '{bd}'::jsonb WHERE name = '{esc(name)}';")
    updates.append(name)

current = None
for r in range(4, ws.max_row + 1):
    p, pack = ws.cell(r, 1).value, ws.cell(r, 2).value
    inv = num(ws.cell(r, 3).value)
    if p and clean(p): current = clean(p)
    if inv is None or pack is None or not current: continue
    pack = clean(pack)
    deal_raw, m_raw = ws.cell(r, 9).value, ws.cell(r, 13).value
    deal = clean(deal_raw) if isinstance(deal_raw, str) and "+" in str(deal_raw) else None
    dealer_deal = clean(m_raw) if isinstance(m_raw, str) and "+" in str(m_raw) else None
    name = db_name(current, pack)

    if 4 <= r <= 34:          # standard cascade
        capital, srp = num(ws.cell(r, 8).value), num(ws.cell(r, 10).value)
        bd = {"model": "standard", "invoice": inv, "less": [0.2, 0.1, 0.2], "vat": "add_12",
              "srp_markup": 1.4, "packaging": pack, "dealer_deal": dealer_deal, "sheet_row": r}
        upsert(name, capital, srp, deal, 0.15 if srp else 0, 0.05 if srp else 0, bd, pack)
    elif 37 <= r <= 40:       # wheat germ
        capital, srp = num(ws.cell(r, 8).value), num(ws.cell(r, 10).value)
        pickup = num(ws.cell(r, 7).value) or 0
        bd = {"model": "wheatgerm", "invoice": inv, "less": ["10%", "5%", "20%"],
              "pickup_pesos": pickup, "vat": "none", "packaging": pack, "sheet_row": r}
        upsert(name, capital, srp, None, 0.15 if srp else 0, 0.05 if srp else 0, bd, pack)
    elif 45 <= r <= 58:       # SCM cascade (Robophos included — it has SRP)
        capital, srp = num(ws.cell(r, 8).value), num(ws.cell(r, 10).value)
        bd = {"model": "scm", "invoice": inv, "less": [0.1, 0.2, 0.05],
              "labels": ["Volume discount 10%", "Pick-up 20%", "PBD 5%"], "vat": "add_12",
              "packaging": pack, "sheet_row": r}
        if srp is not None: bd["srp_markup"] = 1.4
        upsert(name, capital, srp, None, 0.15 if srp else 0, 0.05 if srp else 0, bd, pack)
    elif 63 <= r <= 67:       # dealers discount (peso cascade)
        capital, srp = num(ws.cell(r, 8).value), num(ws.cell(r, 10).value)
        vol, pd_, pbd = num(ws.cell(r, 4).value), num(ws.cell(r, 5).value), num(ws.cell(r, 6).value)
        bd = {"model": "dealer", "invoice": inv, "volume": vol, "pd": pd_, "pbd_pesos": pbd,
              "vat": "add_12", "srp_markup": 1.4, "packaging": pack, "sheet_row": r}
        upsert(name, capital, srp, None, 0.15, 0.05, bd, pack)

# ---- merge duplicates into the STOCKED Pet Supplies items, then drop the dupes ----
MERGES = [  # stocked item  <=  duplicate catalog item created from the dealer section
    ("Top B+ Vitamins 120ml",     "Topb + Multivitamins 120ml x 18bottles"),
    ("Top B+ Vitamins 60ml",      "Topb + Multivitamins 60ml x 36bottles"),
    ("Shampooch 300ml bt 12/box", "Shampooch 300ml x 12bottles"),
    ("Shampooch Sachet",          "Shampooch 15ml x 25sachets x 16box"),
]
for keep, dupe in MERGES:
    sql.append(f"""UPDATE items k SET
      cost = d.cost, sales_price = d.sales_price, deal = d.deal,
      outright_rate = d.outright_rate, cod_rate = d.cod_rate,
      packaging = d.packaging, price_breakdown = d.price_breakdown
    FROM items d WHERE k.name = '{esc(keep)}' AND d.name = '{esc(dupe)}';""")
    sql.append(f"DELETE FROM items WHERE name = '{esc(dupe)}' AND NOT EXISTS "
               f"(SELECT 1 FROM sale_items si WHERE si.item_id = items.id) AND NOT EXISTS "
               f"(SELECT 1 FROM purchases pu WHERE pu.item_id = items.id) AND NOT EXISTS "
               f"(SELECT 1 FROM manual_inventory mi WHERE mi.item_id = items.id);")

names_list = ",".join(f"('{esc(n)}')" for n in updates)
sql.append(f"SELECT v.name AS unmatched FROM (VALUES {names_list}) v(name) "
           f"LEFT JOIN items i ON i.name = v.name WHERE i.id IS NULL;")
sql.append("COMMIT;")
sql.append("SELECT 'robichem items: '||count(*) FROM items WHERE category='Robichem';")
sql.append("SELECT 'total items: '||count(*) FROM items;")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))
print(f"rows re-extracted: {len(updates)}; wrote {OUT}")
