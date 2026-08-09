"""
Full catalog reconciliation against "URC net Capital .xlsx".
- Extracts EVERY RobiChem row (product carried forward across packaging variants).
- Updates existing items with packaging + uom; inserts missing catalog items (stock 0).
- Adds uom to all items by rules (bag/box/bottle/sachet/jar/...).
Generates catalog_update.sql.
"""
import openpyxl, json, os, re

SRC = r"C:\Users\PC\Downloads\URC net Capital .xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "catalog_update.sql")

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["RobiChem"]

def num(v):
    try:
        f = float(v)
        return round(f, 2)
    except (TypeError, ValueError):
        return None

def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip()

def uom_for(pack):
    p = pack.lower()
    if "box" in p: return "box"
    if "kg" in p: return "jar"
    if "liter" in p or p.endswith("l") and "ml" not in p: return "bottle"
    if "ml" in p: return "bottle"
    if "dose" in p or "ds" in p: return "bottle"
    if p.endswith("g.") or p.endswith("g"): return "sachet"
    if "disp" in p: return "set"
    return "piece"

# --- walk RobiChem rows, carrying the product name forward across variants ---
rows = []
current = None
for r in range(4, ws.max_row + 1):
    name, pack = ws.cell(r, 1).value, ws.cell(r, 2).value
    inv = num(ws.cell(r, 3).value)
    if name and clean(name) and inv is None and pack is None:
        continue  # section header text
    if inv is None:
        continue
    if name and clean(name):
        current = clean(name)
    if not current or pack is None:
        continue
    pack = clean(pack)
    capital = num(ws.cell(r, 8).value)          # NET PRICE + 12% vat
    deal_raw = ws.cell(r, 9).value
    deal = clean(deal_raw) if isinstance(deal_raw, str) and "+" in str(deal_raw) else None
    srp = num(ws.cell(r, 10).value)
    if capital is None:
        continue  # broken row
    rows.append({"product": current, "packaging": pack, "invoice": inv,
                 "capital": capital, "srp": srp, "deal": deal, "row": r})

print(f"RobiChem catalog rows extracted: {len(rows)}")

# --- map sheet rows to EXISTING db item names (product, packaging) -> db name ---
EXISTING = {
    ("COCCIBUSTER", "5g."): "Coccibuster",
    ("LEVOMAX", "5g."): "Levomax",
    ("SPECTRUM", "5g."): "Spectrum (96/box)",
    ("SPECTRUM PLUS", "5g."): "Spectrum Plus (96/box)",
    ("ROBISTREP VK Powder", "5g."): "Robistrep Vk 5Gm",
    ("WORM BUSTER", "5g."): "Wormbuster Single Dose 5Gm",
    ("ROBI LA Inj.", "100 ml"): "Robi L.A inj 100ml",
    ("ROBIPENSTREP P", "10 ds w/ diluent"): "Robipenstrep P 10dose/Diluent",
    ("ROBIPENSTREP P", "Single dose"): "Robipenstrep P Single dose bt",
    ("IRON - D Inj.", "100 ml"): "Iron - D inj",
    ("ROBICOMJECT Injectible", "100 ml"): "Robicomject inj 100ml",
    ("TRIPULAC Pig Doser", "2 x 100ml (disp.)"): "Tripulac Pig Doser 2x1 set",
    ("WHEAT GERM /BOX", "300g"): "Wheatgerm 300Gm x 12/box",
}
def existing_key(row):
    for (p, k), dbname in EXISTING.items():
        if clean(p).lower() == row["product"].lower() and clean(k).lower() == row["packaging"].lower():
            return dbname
    return None

def title_name(row):
    # readable catalog name: Product + packaging
    p = row["product"]
    p = re.sub(r"\s+", " ", p).strip().title()
    p = p.replace("Scm", "SCM").replace("Vk", "VK").replace("La", "LA").replace("Ii", "II")
    return f"{p} {row['packaging']}"

sql = ["BEGIN;",
"""ALTER TABLE items
  ADD COLUMN IF NOT EXISTS packaging text,
  ADD COLUMN IF NOT EXISTS uom text;"""]

matched, inserted = 0, 0
for row in rows:
    bd = json.dumps({"invoice": row["invoice"], "less": [0.2, 0.1, 0.2], "vat": "add_12",
                     "srp_markup": 1.4, "packaging": row["packaging"],
                     "sheet_row": row["row"]}).replace("'", "''")
    pack = row["packaging"].replace("'", "''")
    u = uom_for(row["packaging"])
    deal_sql = f"'{row['deal']}'" if row["deal"] else "NULL"
    srp_sql = row["srp"] if row["srp"] is not None else "NULL"
    dbname = existing_key(row)
    if dbname:
        n = dbname.replace("'", "''")
        # special: (96/box) items are boxes of 96 sachets
        p2 = f"{pack} x 96/box" if "(96/box)" in dbname else ("300g x 12/box" if "12/box" in dbname else pack)
        u2 = "box" if "box" in p2 else u
        sql.append(
            f"UPDATE items SET packaging = '{p2}', uom = '{u2}', "
            f"cost = {row['capital']}, sales_price = COALESCE(sales_price, {srp_sql}), "
            f"deal = COALESCE(deal, {deal_sql}), price_breakdown = '{bd}'::jsonb "
            f"WHERE name = '{n}';")
        matched += 1
    else:
        n = title_name(row).replace("'", "''")
        outright = 0.15 if row["srp"] is not None else 0
        cod = 0.05 if row["srp"] is not None else 0
        sql.append(
            "INSERT INTO items (name, category, type, packaging, uom, cost, sales_price, deal, "
            "outright_rate, cod_rate, price_breakdown, notes) VALUES ("
            f"'{n}', 'Robichem', 'Medicine', '{pack}', '{u}', {row['capital']}, {srp_sql}, {deal_sql}, "
            f"{outright}, {cod}, '{bd}'::jsonb, 'From URC catalog — not yet stocked') "
            "ON CONFLICT (name) DO NOTHING;")
        inserted += 1

# --- uom rules for the rest of the catalog ---
sql.append("""
UPDATE items SET uom = 'bag'    WHERE uom IS NULL AND (name ILIKE '%(50KG)%' OR name ILIKE '%(25KG)%' OR name ILIKE '%(20KG)%' OR name ILIKE '%(5KG)%' OR name ILIKE '%(25x1kg)%');
UPDATE items SET uom = 'box'    WHERE uom IS NULL AND (name ILIKE '%box%' OR name ILIKE '%x12%' OR name ILIKE '%x10%' OR name ILIKE '%x 12%' OR notes ILIKE '%Unit: box%');
UPDATE items SET uom = 'carton' WHERE uom IS NULL AND notes ILIKE '%Unit: carton%';
UPDATE items SET uom = 'bundle' WHERE uom IS NULL AND name ILIKE '%(10Lx3)%';
UPDATE items SET uom = 'piece'  WHERE uom IS NULL;""")

sql.append("COMMIT;")
sql.append("SELECT category, count(*) AS items FROM items GROUP BY category ORDER BY 1;")
sql.append("SELECT 'total: '||count(*) FROM items;")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))
print(f"matched existing: {matched}, new catalog items: {inserted}")
print("wrote", OUT)
