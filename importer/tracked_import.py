"""
Import ONLY the DR-tracked (blue) Jerome Racaza / Fegen Gines invoices from
"Upon Checking of Elishen URC Sales until Jul 20.xlsx".
Blue-tracked & not orange-flagged: rows 9, 28, 40, 48, 66 (all PAID).
Orange (no DR) and unmarked invoices stay excluded.
Duplicate-guarded: every insert is skipped if customer+date+total already exists.
"""
import openpyxl, re, os, datetime
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
# reuse the proven product-mapping + qty helpers from the masterlist importer
src = open(os.path.join(HERE, "masterlist_import.py"), encoding="utf-8").read()
exec(src.split("# ---- parse invoices ----")[0])   # defs: norm, esc, qty_of, num, map_item

CHK = r"C:\Users\PC\Downloads\Upon Checking of Elishen URC Sales until Jul 20.xlsx"
ws = openpyxl.load_workbook(CHK, data_only=True)["MASTERLIST"]

TRACKED_ROWS = [9, 28, 40, 48, 66]      # invoice start rows marked tracked (blue), not orange
OUT = os.path.join(HERE, "tracked.sql")

def read_invoice(start):
    date = ws.cell(start, 1).value.date().isoformat()
    cust = re.sub(r"\s+", " ", str(ws.cell(start, 2).value)).strip()
    addr = ws.cell(start, 3).value
    disc = num(ws.cell(start, 7).value) or 0
    total = num(ws.cell(start, 8).value)
    lines, r = [], start
    while r <= ws.max_row:
        if r > start and isinstance(ws.cell(r, 1).value, datetime.datetime):
            break
        prod = ws.cell(r, 4).value
        if prod and str(prod).strip():
            item = map_item(prod)
            if item is None:
                raise SystemExit(f"UNMATCHED product at R{r}: {prod}")
            lines.append({"item": item, "qty": qty_of(ws.cell(r, 5).value) if ws.cell(r, 5).value is not None else 1.0,
                          "amount": num(ws.cell(r, 6).value) or 0})
        r += 1
    return {"date": date, "cust": cust, "addr": re.sub(r"\s+", " ", str(addr)).strip() if addr else None,
            "disc": disc, "total": total, "lines": lines}

sql = ["BEGIN;"]
offsets = defaultdict(float)
n = 77
for start in TRACKED_ROWS:
    inv = read_invoice(start)
    sn = f"ML-{n:03d}"; n += 1
    subtotal = sum(l["amount"] for l in inv["lines"])
    guard = (f"NOT EXISTS (SELECT 1 FROM sales x WHERE UPPER(TRIM(x.customer)) = UPPER('{esc(inv['cust'])}') "
             f"AND x.date = '{inv['date']}' AND x.total = {inv['total']})")
    sql.append(
        "INSERT INTO sales (sales_no, date, customer, store_farm, subtotal, tax_pct, tax_amount, "
        "discount_pct, discount, total, amount_paid, status) "
        f"SELECT '{sn}', '{inv['date']}', '{esc(inv['cust'])}', "
        f"{('null' if not inv['addr'] else chr(39)+esc(inv['addr'])+chr(39))}, "
        f"{round(subtotal,2)}, 0, 0, 0, {inv['disc']}, {inv['total']}, {inv['total']}, 'Completed' "
        f"WHERE {guard};")
    for l in inv["lines"]:
        up = round(l["amount"] / l["qty"], 4) if l["qty"] else 0
        sql.append(
            "INSERT INTO sale_items (sale_id, item_id, qty, unit_price, total_price) "
            f"SELECT s.id, i.id, {l['qty']}, {up}, {l['amount']} FROM sales s, items i "
            f"WHERE s.sales_no = '{sn}' AND i.name = '{esc(l['item'])}' "
            f"AND NOT EXISTS (SELECT 1 FROM sale_items e WHERE e.sale_id = s.id AND e.item_id = i.id AND e.qty = {l['qty']});")
        offsets[(l["item"], sn)] = offsets.get((l["item"], sn), 0) + l["qty"]
    sql.append(
        "INSERT INTO payments (sale_id, date, amount, account_id, notes) "
        f"SELECT s.id, '{inv['date']}', {inv['total']}, NULL, "
        f"'Checked masterlist — DR tracked (blue), marked PAID' FROM sales s WHERE s.sales_no = '{sn}' "
        f"AND NOT EXISTS (SELECT 1 FROM payments e WHERE e.sale_id = s.id);")

# stock-neutral offsets (these sales predate the 7/29 count), guarded per source invoice
for (item, sn), q in sorted(offsets.items()):
    sql.append(
        "INSERT INTO manual_inventory (date, batch_no, item_id, qty, notes) "
        f"SELECT '2026-07-29', 'PRE-STOCKTAKE-OFFSET', i.id, {round(q,3)}, "
        f"'Offset for {sn}: tracked sale predates the 7/29 physical count' "
        f"FROM items i WHERE i.name = '{esc(item)}' "
        f"AND EXISTS (SELECT 1 FROM sales WHERE sales_no = '{sn}') "
        f"AND NOT EXISTS (SELECT 1 FROM manual_inventory e WHERE e.item_id = i.id AND e.notes LIKE 'Offset for {sn}%');")

sql.append("COMMIT;")
sql.append("SELECT sales_no, date, customer, total, amount_paid FROM sales WHERE sales_no >= 'ML-077' ORDER BY sales_no;")
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql))
print("wrote", OUT)
